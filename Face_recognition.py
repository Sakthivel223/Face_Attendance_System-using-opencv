import os
import cv2
import numpy as np
import openpyxl
import face_recognition
import json
from datetime import datetime






KNOWN_FACES_FILE = "known_faces.json"
UNKNOWN_LOG_FILE = "Unknown Persons.xlsx"

def log_attendance(name, is_known=True):
    """Log attendance for known and unknown persons locally.
    Returns True if successful, False otherwise."""
    now = datetime.now()
    # Prepare data for Excel file
    data = []
    month_year = now.strftime("%b %Y")
    file_name = f"{month_year}.xlsx" if is_known else UNKNOWN_LOG_FILE
    date_today = now.strftime("%Y-%m-%d")
    time_now = now.strftime("%H:%M:%S")







    month_year = now.strftime("%b %Y")
    file_name = f"{month_year}.xlsx" if is_known else UNKNOWN_LOG_FILE
    date_today = now.strftime("%Y-%m-%d")
    time_now = now.strftime("%H:%M:%S")
    
    try:
        # Try to load existing workbook
        try:
            wb = openpyxl.load_workbook(file_name)
            ws = wb.active
        except FileNotFoundError:
            # Create new workbook if it doesn't exist
            wb = openpyxl.Workbook()
            ws = wb.active
            if is_known:
                ws.append(["Name", "Date", "Entry Time", "Exit Time"])
            else:
                ws.append(["Unknown Name", "Date", "Entry Time"])

        # Check for existing entry
        entry_found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == name and row[1] == date_today:
                entry_found = True
                if is_known and row[3] == "-":
                    # Update exit time for known person
                    ws.cell(row=ws.max_row, column=4, value=time_now)
                break

        if not entry_found:
            # Add new entry
            if is_known:
                ws.append([name, date_today, time_now, "-"])
            else:
                ws.append([name, date_today, time_now])

        # Save the workbook
        wb.save(file_name)

        print(f"{name}'s attendance logged successfully in {file_name}")
        return True

    except Exception as e:
        print(f"Error logging attendance for {name}: {e}")
        return False




def recognize_faces_in_camera():
    """Recognize faces in real-time camera feed and log attendance."""
    try:
        with open(KNOWN_FACES_FILE, "r") as f:
            known_faces = json.load(f)
    except Exception as e:
        print(f"Error loading known faces: {e}")
        return

    
    known_face_encodings = [np.array(enc) for enc in known_faces.values()]
    known_face_names = list(known_faces.keys())
    
    video_capture = cv2.VideoCapture(0)
    
    if not video_capture.isOpened():
        print("Error: Could not open camera.")
        return
    
    print("Camera started. Press 'q' to exit.")
    while True:
        ret, frame = video_capture.read()
        if not ret:
            print("Error: Unable to capture frame.")
            break
        
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
        
        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
            # Get face distances and find the best match
            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances)
            
            # Only consider it a match if the distance is below our threshold
            if face_distances[best_match_index] < 0.5:
                name = known_face_names[best_match_index]
                color = (0, 255, 0)  # Green for known faces
            else:
                name = "Unknown"
                color = (0, 0, 255)  # Red for unknown faces
                if log_attendance(name, is_known=False):
                    print(f"Logged unknown person's attendance")


            
            cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
            cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)
        
        cv2.imshow('Face Recognition Attendance System', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    
    video_capture.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    recognize_faces_in_camera()
